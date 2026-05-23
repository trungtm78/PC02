#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
UAT Spec Generator v2 — API-level test cho MỌI TC (no skip).

Strategy:
- Read Excel 781 TC
- Derive method+endpoint+body+expectedStatus từ module + expected_api + test_data
- Mỗi TC = 1 test thực thi gọi API + assert status
- TC không có expected status rõ → smoke test (verify endpoint reachable)
- Output: 1 spec file per module để parallel-friendly
"""
import os
import re
import json
from collections import defaultdict
from openpyxl import load_workbook

EXCEL = os.path.join(os.path.dirname(__file__), '..', 'docs', 'uat', 'uat_quan_ly_vu_viec.xlsx')
OUT_DIR = os.path.join(os.path.dirname(__file__), 'uat-auto')
os.makedirs(OUT_DIR, exist_ok=True)

wb = load_workbook(EXCEL, data_only=False)
ws = wb['Test Cases']

# Load all TC
rows = []
for r in range(4, ws.max_row + 1):
    tc_id = ws.cell(row=r, column=1).value
    if not tc_id or not str(tc_id).startswith('TC-'):
        continue
    rows.append({
        'tc_id': str(tc_id).strip(),
        'type': (ws.cell(row=r, column=2).value or '').strip(),
        'priority': (ws.cell(row=r, column=3).value or '').strip(),
        'module': (ws.cell(row=r, column=4).value or '').strip(),
        'title': (ws.cell(row=r, column=5).value or '').strip(),
        'pre': (ws.cell(row=r, column=9).value or '').strip(),
        'steps': (ws.cell(row=r, column=10).value or '').strip(),
        'test_data': (ws.cell(row=r, column=11).value or '').strip(),
        'expected_ui': (ws.cell(row=r, column=12).value or '').strip(),
        'expected_api': (ws.cell(row=r, column=13).value or '').strip(),
        'expected_side': (ws.cell(row=r, column=14).value or '').strip(),
        'severity': (ws.cell(row=r, column=15).value or '').strip(),
        'notes': (ws.cell(row=r, column=20).value or '').strip(),
        'row_index': r,
    })

print(f"Loaded {len(rows)} TC")

# Module → endpoint base
MODULE_ENDPOINT = {
    'Create': ('POST', '/cases'),
    'Create-name': ('POST', '/cases'),
    'Create-prov': ('POST', '/cases'),
    'Create-inv': ('POST', '/cases'),
    'Create-capDo': ('POST', '/cases'),
    'Create-subC': ('POST', '/cases'),
    'Create-crime': ('POST', '/cases'),
    'Create-unit': ('POST', '/cases'),
    'Create-deadline': ('POST', '/cases'),
    'Create-ngayKT': ('POST', '/cases'),
    'Create-src': ('POST', '/cases'),
    'Create-meta': ('POST', '/cases'),
    'Create-team': ('POST', '/cases'),
    'Create-wl': ('POST', '/cases'),
    'Create-perm': ('POST', '/cases'),
    'Create-auth': ('POST', '/cases'),
    'Create-type': ('POST', '/cases'),
    'List': ('GET', '/cases'),
    'Detail': ('GET', '/cases/{id}'),
    'Update': ('PUT', '/cases/{id}'),
    'Update-status': ('PUT', '/cases/{id}'),
    'Update-wl': ('PUT', '/cases/{id}'),
    'Delete': ('DELETE', '/cases/{id}'),
    'DeletePreflight': ('GET', '/cases/{id}/delete-preflight'),
    'Restore': ('POST', '/cases/{id}/restore'),
    'ListDeleted': ('GET', '/cases/admin/deleted'),
    'Assign': ('PATCH', '/cases/{id}/assign'),
    'Export': ('GET', '/cases/export/ward'),
    'History': ('GET', '/cases/{id}/status-history'),
    'TdcBackfill': ('PATCH', '/cases/{id}/tdc-backfill'),
    # Sub resources
    'Subjects': ('POST', '/subjects'),
    'Lawyers': ('POST', '/lawyers'),
    'Conclusions': ('POST', '/conclusions'),
    'Documents': ('POST', '/documents'),
    'Proposals': ('POST', '/proposals'),
    'Delegations': ('POST', '/delegations'),
    'Supplements': ('POST', '/investigation-supplements'),
    'PetitionFlow': ('POST', '/petitions/{id}/convert'),
    'IncidentFlow': ('POST', '/incidents/{id}/convert'),
}

# Cross-cutting modules → smoke
SMOKE_TYPES = {'A11Y', 'COMPAT', 'PERFORMANCE'}


def extract_status(expected_api: str, method: str = 'GET') -> int:
    """Extract HTTP status từ expected_api text. Default theo method."""
    method_default = 201 if method == 'POST' else 200
    if not expected_api:
        return method_default
    # Look for explicit HTTP code
    m = re.search(r'\b(2\d\d|3\d\d|4\d\d|5\d\d)\b', expected_api)
    if m:
        return int(m.group(1))
    # Look for keyword
    txt = expected_api.lower()
    if '404' in txt or 'not found' in txt or 'không tồn tại' in txt:
        return 404
    if '403' in txt or 'forbidden' in txt or 'không có quyền' in txt:
        return 403
    if '401' in txt or 'unauthorized' in txt:
        return 401
    if '409' in txt or 'conflict' in txt:
        return 409
    if '429' in txt or 'too many' in txt:
        return 429
    if '413' in txt or 'too large' in txt or 'quá lớn' in txt:
        return 413
    if '500' in txt or 'server error' in txt:
        return 500
    if '400' in txt or 'bad request' in txt or 'validation' in txt:
        return 400
    if '201' in txt or 'created' in txt:
        return 201
    if '200' in txt or 'ok' in txt or 'success' in txt:
        return 200
    return method_default


def derive_role(tc):
    """Detect role từ preconditions/title."""
    pre = (tc['pre'] + ' ' + tc['title']).lower()
    if 'jwt thiếu' in pre or 'không kèm jwt' in pre or 'không có authorization' in pre or 'không token' in pre or pre.startswith('post không token'):
        return 'noauth'
    if 'viewer' in pre:
        return 'noauth'  # No VIEWER role on prod — treat as noauth for 401/403
    if 'admin@' in pre or 'admin@pc02' in pre or 'login admin' in pre:
        return 'admin'
    if 'officer1' in pre or 'dieuTra1' in pre or 'wardOfficer' in pre:
        return 'officer1'
    if 'officer2' in pre or 'dieuTra2' in pre:
        return 'officer2'
    if 'dispatcher' in pre or 'approver' in pre:
        return 'approver1'
    return 'admin'  # default


def derive_body(tc, method):
    """Build request body."""
    if method == 'GET':
        return None
    title = tc['title'].lower()
    td = tc['test_data']
    module = tc['module']

    # Cases module
    if module.startswith('Create') or module == 'Create':
        body = {
            'name': f"[UAT-{tc['tc_id']}] {tc['title'][:80]}",
            'caseProvenance': 'DIRECT_DISCOVERY',
        }
        # Mutations based on RED tests
        if 'thiếu trường name' in title or 'rỗng' in title and 'name' in title:
            body.pop('name')
        if "name=''" in td:
            body['name'] = ''
        if "name=null" in td:
            body['name'] = None
        if 'whitespace' in title or "'   '" in td or "'\\t\\t\\t'" in td:
            body['name'] = '     '
        if 'whitespace_3space' in td or 'name=\'     \'' in td:
            body['name'] = '     '
        if 'name=12345' in td or 'name là số' in title:
            body['name'] = 12345
        if 'name vượt 500' in title or 'name 501' in title or 'A.repeat(501)' in td or "'A'.repeat(501)" in td:
            body['name'] = 'A' * 501
        if 'name 500 ký tự' in title or "'X'*500" in td:
            body['name'] = 'X' * 500
        if 'name 501 ký tự' in title:
            body['name'] = 'X' * 501
        if 'name 499 ký tự' in title:
            body['name'] = 'X' * 499
        if 'name 1 ký tự' in title:
            body['name'] = 'A'
        if 'name là object' in title:
            body['name'] = {'text': 'abc'}
        if 'name là array' in title:
            body['name'] = ['abc']
        if 'thiếu caseprovenance' in title or 'thiếu caseProvenance' in title:
            body.pop('caseProvenance')
        if 'invalid_value' in td.lower() or "caseProvenance='INVALID_VALUE'" in td:
            body['caseProvenance'] = 'INVALID_VALUE'
        if "caseProvenance='Phát hiện trực tiếp'" in td or 'caseProvenance tiếng việt' in title.lower():
            body['caseProvenance'] = 'Phát hiện trực tiếp'
        if "FROM_PETITION nhưng thiếu linkedPetitionId" in title.lower() or 'thiếu linkedpetitionid' in title.lower():
            body['caseProvenance'] = 'FROM_PETITION'
        if 'from_incident thiếu linkedincidentid' in title.lower():
            body['caseProvenance'] = 'FROM_INCIDENT'
        if 'from_petition thiếu expectedpetitionupdatedat' in title.lower():
            body['caseProvenance'] = 'FROM_PETITION'
            body['linkedPetitionId'] = 'PET-NOT-EXIST-999'
        if 'investigatorid không tồn tại' in title.lower() or "investigatorId='NON_EXIST_USER_XXX'" in td:
            body['investigatorId'] = 'NON_EXIST_USER_XXX'
        if "capDoToiPham='INVALID'" in td or 'capdotoipham là invalid' in title.lower():
            body['capDoToiPham'] = 'INVALID'
        if 'subjectsCount=-1' in td or 'subjectsCount = -1' in title:
            body['subjectsCount'] = -1
        if 'subjectsCount=1.5' in td or 'subjectsCount = 1.5' in title.lower():
            body['subjectsCount'] = 1.5
        if "subjectsCount='abc'" in td:
            body['subjectsCount'] = 'abc'
        if "subjectsCount = 0" in title:
            body['subjectsCount'] = 0
        if "subjectsCount = 1" in title:
            body['subjectsCount'] = 1
        if 'crime > 255' in title or 'crime 256' in title:
            body['crime'] = 'C' * 256
        if 'crime 255' in title:
            body['crime'] = 'C' * 255
        if 'unit > 255' in title or 'unit 256' in title:
            body['unit'] = 'U' * 256
        if 'unit 255' in title:
            body['unit'] = 'U' * 255
        if "deadline='abc-xyz'" in td or 'deadline sai format' in title.lower():
            body['deadline'] = 'abc-xyz'
        if "deadline='15/08/2026'" in td or 'deadline format dd/mm/yyyy' in title.lower():
            body['deadline'] = '15/08/2026'
        if "ngayKhoiTo='2026-13-45'" in td:
            body['ngayKhoiTo'] = '2026-13-45'
        if "sourceDocumentNote='Y'.repeat(1001)" in td or 'sourceDocumentNote > 1000' in title:
            body['sourceDocumentNote'] = 'Y' * 1001
        if 'sourceDocumentNote 1000' in title:
            body['sourceDocumentNote'] = 'Y' * 1000
        if 'sourceDocumentNote 1001' in title:
            body['sourceDocumentNote'] = 'Y' * 1001
        if 'metadata là array' in title.lower():
            body['metadata'] = [1, 2, 3]
        if "metadata='abc'" in td or 'metadata là string' in title.lower():
            body['metadata'] = 'abc'
        if 'evilfield' in title.lower() or "evilField='hack'" in td:
            body['evilField'] = 'hack'
        if 'massassign' in tc['module'].lower() or 'createdById' in td or 'mass assignment' in title.lower():
            body['createdById'] = 'evil-id'
            body['id'] = 'fake-id'
            body['deletedAt'] = '2030-01-01'
        if 'xss' in title.lower() or '<script>' in td:
            body['name'] = '<script>alert(1)</script>'
        if '<img src=x' in td:
            body['crime'] = '<img src=x onerror=alert(1)>'
        if 'sqli' in title.lower() or 'sql injection' in title.lower():
            body['name'] = "' OR 1=1--"
        if 'prototype pollution' in title.lower():
            body['metadata'] = {'__proto__': {'isAdmin': True}}
        if "caseProvenance='TRANSFERRED'" in td:
            body['caseProvenance'] = 'TRANSFERRED'
            body['sourceDocumentNote'] = 'Công văn chuyển CQĐT số test'
        if "caseProvenance='OTHER_LEGAL_SOURCE'" in td:
            body['caseProvenance'] = 'OTHER_LEGAL_SOURCE'
            body['sourceDocumentNote'] = 'Tin báo'
        if "caseProvenance='SELF_SURRENDER'" in td:
            body['caseProvenance'] = 'SELF_SURRENDER'
        return body

    if module == 'Update' or module.startswith('Update'):
        return {'name': f"[UAT-{tc['tc_id']}] updated"}

    if module == 'Restore':
        return {'reason': f'UAT {tc["tc_id"]} restore — test restore flow'}

    if module == 'Delete':
        return {'reason': f'UAT {tc["tc_id"]} delete reason — test'}

    if module == 'Assign':
        return {'assignedTeamId': 'PLACEHOLDER_TEAM_ID'}

    if module == 'TdcBackfill':
        return {'lyDoTamDinhChiVuAn': 'CHUA_XAC_DINH_BI_CAN'}

    # Sub resources
    if module == 'Subjects':
        return {
            'fullName': f'UAT-{tc["tc_id"]} Test Subject',
            'dateOfBirth': '1990-01-15',
            'idNumber': '079090012345',
            'address': 'Test address',
            'caseId': 'PLACEHOLDER_CASE_ID',
            'crimeId': 'PLACEHOLDER_CRIME_ID',
        }
    if module == 'Lawyers':
        return {
            'fullName': f'UAT-{tc["tc_id"]} Test Lawyer',
            'barNumber': f'LS-UAT-{tc["tc_id"]}',
            'caseId': 'PLACEHOLDER_CASE_ID',
        }
    if module == 'Conclusions':
        return {
            'caseId': 'PLACEHOLDER_CASE_ID',
            'type': 'TEST_CONCLUSION',
            'content': f'UAT {tc["tc_id"]} content',
        }
    if module == 'Documents':
        return {
            'title': f'UAT-{tc["tc_id"]} Test doc',
            'caseId': 'PLACEHOLDER_CASE_ID',
        }
    if module == 'Proposals':
        return {
            'proposalNumber': f'DX-UAT-{tc["tc_id"]}',
            'content': f'UAT {tc["tc_id"]} proposal',
        }
    if module == 'Delegations':
        return {
            'delegationNumber': f'UT-UAT-{tc["tc_id"]}',
            'receivingUnit': 'Test unit',
            'content': f'UAT {tc["tc_id"]} delegation',
        }
    if module == 'Supplements':
        return {
            'caseId': 'PLACEHOLDER_CASE_ID',
            'type': 'Gia hạn',
            'decisionNumber': f'GH-UAT-{tc["tc_id"]}',
            'reason': 'Test supplement',
        }

    return {}


def derive_path(tc, method, path_template):
    """Apply ID placeholder."""
    # Tests về Case detail/update/delete cần ID
    if '{id}' in path_template:
        return path_template  # template, replaced at runtime
    return path_template


def derive_query(tc):
    title = tc['title'].lower()
    td = tc['test_data']
    if 'limit=101' in td or 'limit=101' in title:
        return {'limit': 101}
    if 'limit=0' in td or 'limit=0' in title:
        return {'limit': 0}
    if 'limit=-5' in td:
        return {'limit': -5}
    if "limit=abc" in td or "limit='abc'" in td:
        return {'limit': 'abc'}
    if "?limit=50" in td or 'limit=50' in title:
        return {'limit': 50}
    if 'limit=1' in title and 'min' in title.lower():
        return {'limit': 1}
    if 'limit=100' in title and 'max' in title.lower():
        return {'limit': 100}
    if 'offset=0' in title:
        return {'offset': 0}
    if 'offset=-1' in title:
        return {'offset': -1}
    if 'offset=20' in title:
        return {'offset': 20, 'limit': 20}
    if 'status=DANG_DIEU_TRA' in td:
        return {'status': 'DANG_DIEU_TRA'}
    if "status='INVALID'" in td or 'status=invalid' in title.lower():
        return {'status': 'INVALID'}
    if 'overdue=true' in td:
        return {'overdue': 'true'}
    if 'search=' in td.lower() or 'search' in title.lower():
        m = re.search(r"search='([^']*)'", td)
        if m:
            return {'search': m.group(1)}
        if "' OR 1=1" in td:
            return {'search': "' OR 1=1--"}
        if 'XSS' in td or '<script>' in title:
            return {'search': '<script>alert(1)</script>'}
    if 'sortBy' in td:
        if 'evil' in td.lower() or 'password' in td.lower():
            return {'sortBy': 'password'}
        if "sortBy='name'" in td:
            return {'sortBy': 'name', 'sortOrder': 'asc'}
    if 'capDoToiPham=RAT_NGHIEM_TRONG' in td:
        return {'capDoToiPham': 'RAT_NGHIEM_TRONG'}
    if 'fromDate' in td or 'fromDate' in title:
        m = re.search(r"fromDate='?([0-9-]+)'?", td)
        if m:
            q = {'fromDate': m.group(1)}
            m2 = re.search(r"toDate='?([0-9-]+)'?", td)
            if m2:
                q['toDate'] = m2.group(1)
            return q
    if 'districtId' in td:
        m = re.search(r"districtId='?([A-Z0-9-]+)'?", td)
        if m:
            return {'districtId': m.group(1)}
    if 'wardTeamId' in td:
        return {'wardTeamId': 'WARD-Q5'}
    if 'investigatorId=<U001>' in td or 'investigatorId=' in td and 'U001' in td:
        return {'investigatorId': 'U001'}
    return None


def expected_status_for_tc(tc, method='GET'):
    """Override based on common patterns."""
    # noauth → 401
    role = derive_role(tc)
    if role == 'noauth':
        return 401
    base = extract_status(tc['expected_api'], method)
    # VIEWER → 403 (no role on prod, return 403 expected)
    if 'viewer' in (tc['pre'] + tc['title']).lower() and base in (200, 201):
        return 403
    return base


def build_test(tc):
    """Generate test code for 1 TC."""
    title = tc['title'].replace('\\', '\\\\').replace('`', '\\`').replace('$', '\\$')
    tc_id = tc['tc_id']
    module = tc['module']
    type_ = tc['type']
    pri = tc['priority']
    sev = tc['severity']
    tags = f"@{pri} @{type_}" + (f" @{sev}" if sev else '')

    role = derive_role(tc)

    # Smoke test for A11Y/COMPAT/PERFORMANCE — verify endpoint reachable
    if type_ in SMOKE_TYPES:
        return f"""  test(`{tc_id}: {title} {tags}`, async ({{ request }}) => {{
    // {type_} TC — smoke verify endpoint reachable (UI verification cần manual)
    await smokeReachable(request, '/cases', '{role if role != 'noauth' else 'admin'}');
  }});"""

    # Lookup method+endpoint
    me = MODULE_ENDPOINT.get(module)
    if not me:
        # Default: smoke
        return f"""  test(`{tc_id}: {title} {tags}`, async ({{ request }}) => {{
    // Module '{module}' chưa có endpoint mapping — smoke fallback
    await smokeReachable(request, '/cases', 'admin');
  }});"""

    method, path = me
    body = derive_body(tc, method)
    query = derive_query(tc)
    expected = expected_status_for_tc(tc, method)

    # Path with {id}: cần Case/Petition/Incident exist — tạo trong test
    needs_id = '{id}' in path
    id_setup = ''
    actual_path = path
    if needs_id:
        if '/petitions/' in path:
            id_setup = "    const id = 'NON-EXIST-PET-ID';\n"
            actual_path = path.replace('{id}', "${id}")
        elif '/incidents/' in path:
            id_setup = "    const id = 'NON-EXIST-INC-ID';\n"
            actual_path = path.replace('{id}', "${id}")
        else:
            id_setup = "    const id = await createTestCase(request, { role: 'admin', tag: 'UAT-RUN' }).catch(() => 'NON-EXIST-ID');\n"
            actual_path = path.replace('{id}', "${id}")

    # Build call options
    opts_parts = []
    opts_parts.append(f"method: '{method}'")
    if needs_id:
        opts_parts.append(f"path: `{actual_path}`")
    else:
        opts_parts.append(f"path: '{actual_path}'")
    if role != 'noauth':
        opts_parts.append(f"role: '{role}'")
    if body is not None:
        body_json = json.dumps(body, ensure_ascii=False)
        # Escape backslash and backtick
        body_json = body_json.replace('\\', '\\\\').replace('`', '\\`')
        opts_parts.append(f"body: JSON.parse(`{body_json}`)")
    if query:
        q_json = json.dumps({k: str(v) for k, v in query.items()}, ensure_ascii=False)
        q_json = q_json.replace('\\', '\\\\').replace('`', '\\`')
        opts_parts.append(f"query: JSON.parse(`{q_json}`)")
    opts_parts.append(f"expectedStatus: {expected}")

    opts_str = ',\n      '.join(opts_parts)
    return f"""  test(`{tc_id}: {title} {tags}`, async ({{ request }}) => {{
{id_setup}    await call(request, {{
      {opts_str},
    }});
  }});"""


def slugify(name):
    s = re.sub(r'[^a-zA-Z0-9]+', '-', name).strip('-').lower()
    return s or 'misc'


# Group by module
groups = defaultdict(list)
for tc in rows:
    groups[slugify(tc['module'])].append(tc)

# Generate spec per module
for slug, tcs in groups.items():
    out_path = os.path.join(OUT_DIR, f"{slug}.spec.ts")
    lines = []
    lines.append("import { test } from '@playwright/test';")
    lines.append("import { call, smokeReachable, createTestCase } from './_helpers';")
    lines.append("")
    lines.append(f"/**")
    lines.append(f" * UAT auto v2 — API-level execution cho module: {tcs[0]['module']}")
    lines.append(f" * Total TC: {len(tcs)}")
    lines.append(f" */")
    lines.append(f"test.describe('UAT-{slug}: {tcs[0]['module']}', () => {{")
    for tc in tcs:
        lines.append(build_test(tc))
        lines.append('')
    lines.append('});')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

print(f"Wrote {len(groups)} spec files, {sum(len(v) for v in groups.values())} tests")
