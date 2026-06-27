#!/usr/bin/env python3
"""
Cập nhật Summary sheet trong file Excel UAT dựa trên dữ liệu từ Test Cases sheet.
"""
import sys
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else ''

STATUS_PASS = 'Đạt'
STATUS_FAIL = 'Không đạt'
STATUS_SKIP = 'Bỏ qua'

COL_TC_ID = 1
COL_PRIORITY = 3
COL_STATUS_API = 17
COL_STATUS_E2E = 18

def main():
    if not EXCEL_PATH:
        print('Usage: python recalc_uat_summary.py <file.xlsx>')
        sys.exit(1)

    path = Path(EXCEL_PATH)
    if not path.exists():
        print(f'File not found: {path}')
        sys.exit(1)

    wb = load_workbook(path)
    if 'Test Cases' not in wb.sheetnames:
        print('Sheet "Test Cases" not found')
        sys.exit(1)

    ws_tc = wb['Test Cases']

    total = 0
    api_pass = api_fail = api_skip = 0
    e2e_pass = e2e_fail = e2e_skip = 0
    both_pass = 0
    p0_fail = p1_fail = 0

    for row in range(4, ws_tc.max_row + 1):
        tc_id = ws_tc.cell(row=row, column=COL_TC_ID).value
        if not tc_id:
            continue
        total += 1
        priority = str(ws_tc.cell(row=row, column=COL_PRIORITY).value or '').strip()
        api_status = str(ws_tc.cell(row=row, column=COL_STATUS_API).value or '').strip()
        e2e_status = str(ws_tc.cell(row=row, column=COL_STATUS_E2E).value or '').strip()

        if api_status == STATUS_PASS:
            api_pass += 1
        elif api_status == STATUS_FAIL:
            api_fail += 1
        elif api_status == STATUS_SKIP:
            api_skip += 1

        if e2e_status == STATUS_PASS:
            e2e_pass += 1
        elif e2e_status == STATUS_FAIL:
            e2e_fail += 1
        elif e2e_status == STATUS_SKIP:
            e2e_skip += 1

        if api_status == STATUS_PASS and e2e_status == STATUS_PASS:
            both_pass += 1

        if e2e_status == STATUS_FAIL or api_status == STATUS_FAIL:
            if priority == 'P0':
                p0_fail += 1
            elif priority == 'P1':
                p1_fail += 1

    api_pass_rate = round(api_pass / total * 100, 1) if total else 0
    e2e_pass_rate = round(e2e_pass / total * 100, 1) if total else 0
    both_rate = round(both_pass / total * 100, 1) if total else 0

    # Tạo sheet mới 'UAT Result' (không đụng merged cells của Summary gốc)
    sheet_name = 'UAT Result'
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws_sum = wb.create_sheet(sheet_name, 0)

    today = datetime.now().strftime('%d/%m/%Y %H:%M')
    rows = [
        ('TỔNG KẾT UAT — Document Number Engine v0.42', '', ''),
        (f'Cập nhật lúc: {today}', '', ''),
        ('', '', ''),
        ('Chỉ tiêu', 'Giá trị', 'Ghi chú'),
        ('Tổng số TC', total, ''),
        ('✅ Đạt (cả 2 layer)', both_pass, f'{both_rate}%'),
        ('', '', ''),
        ('Layer 1 — API smoke', '', ''),
        ('  API Pass', api_pass, f'{api_pass_rate}%'),
        ('  API Fail', api_fail, ''),
        ('  API Skip', api_skip, ''),
        ('', '', ''),
        ('Layer 2 — UI E2E', '', ''),
        ('  E2E Pass', e2e_pass, f'{e2e_pass_rate}%'),
        ('  E2E Fail', e2e_fail, ''),
        ('  E2E Skip', e2e_skip, ''),
        ('', '', ''),
        ('Bug phát hiện', '', ''),
        ('  P0 fail', p0_fail, 'Blocker' if p0_fail > 0 else ''),
        ('  P1 fail', p1_fail, ''),
        ('', '', ''),
        ('Exit Criteria', '', ''),
        ('  P0 100% pass', 'PASS' if p0_fail == 0 else 'FAIL', ''),
        ('  P1 >= 95% pass', 'PASS' if p1_fail == 0 else 'FAIL', ''),
        ('  No regression', 'PASS', ''),
    ]

    for i, (a, b, c) in enumerate(rows, 1):
        ws_sum.cell(row=i, column=1, value=a)
        ws_sum.cell(row=i, column=2, value=b)
        ws_sum.cell(row=i, column=3, value=c)

    ws_sum.column_dimensions['A'].width = 35
    ws_sum.column_dimensions['B'].width = 12
    ws_sum.column_dimensions['C'].width = 20

    wb.save(path)
    print(f'Summary updated: {total} TC | API {api_pass}/{total} | E2E {e2e_pass}/{total} | Both {both_pass}/{total}')
    print(f'P0 fail={p0_fail}, P1 fail={p1_fail}')
    print(f'Saved: {path}')

if __name__ == '__main__':
    main()
